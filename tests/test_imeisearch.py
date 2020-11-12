import pytest
from selenium import webdriver
import openpyxl

from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

from PageObjects.InventorySearchPage import InventorySearchPage
from PageObjects.LoginPage import LoginPage
from PageObjects.MyAccountPage import MyAccountPage
from utilities.BaseClass import BaseClass


#@pytest.mark.fixtures("setup")
class TestImeiSearch(BaseClass):
    def test_imeisearch(self):
        # Accessing Xl sheet
        workbook = openpyxl.load_workbook(self.excel_Location)
        sheet = workbook.active

        self.driver.implicitly_wait(5)
        #wait = WebDriverWait(self.driver, 5)
        # Login
        log = self.getLogger()
        loginPage = LoginPage(self.driver)

        # Extracting Captcha numbers
        captchatext = loginPage.login_captcha().text
        print(captchatext)
        print(type(captchatext))
        numbers = []
        for letters in captchatext.split():
            if letters.isdigit():
                numbers.append(int(letters))
        result = sum(numbers)
        log.info(result)
        loginPage.Login_captchaResponse().send_keys(result)
        loginPage.Login_submitBtn().click()

        # Myaccount
        myaccountpage = MyAccountPage(self.driver)
        my_account = myaccountpage.MyAccountBtn().text
        log.info("Text received is "+my_account)
        if my_account == "MY ACCOUNT":
            myaccountpage.MyAccountBtn().click()
            self.verifyLinkPresence("Please use this.")
            myaccountpage.MyAccount_inventorylink().click()
            currenturl = self.driver.current_url
            print(currenturl)
            Search_page = self.driver.window_handles[1] #accessing 2nd page
            print(Search_page)
            self.driver.switch_to.window(Search_page)
            Search_url = self.driver.current_url #to get current url
            log.info(Search_url)
            if Search_url == "http://intranet.sonimtech.com/inventory-search":
                maxrow = len([row for row in sheet if not all([cell.value == None for cell in row])])  # To get max row
                for i in range(2, maxrow + 1):  # iterating each row from excel
                    imei = sheet.cell(row=i, column=1).value
                    inventorySearchpage = InventorySearchPage(self.driver)
                    inventorySearchpage.InventoryImeiSearch().send_keys(imei)
                    log.info("Fetching IMEI number from excel")
                    inventorySearchpage.InventorySearch_submitBtn().click()
                    self.verifyElementPresence("//tr[1]/td[1]/a")
                    imeinumber = inventorySearchpage.InventorySearch_imeiNo().text
                    log.info(imeinumber)
                    self.verifyElementPresence("//tr[1]/td[2]")
                    employee = inventorySearchpage.InventorySearch_name().text
                    self.verifyElementPresence("//tr[1]/td[5]")
                    model = inventorySearchpage.InventorySearch_modelName().text
                    status = inventorySearchpage.InventorySearch_Status().text
                    # writing employee, model and status into  the sheet
                    sheet.cell(row=i, column=2).value = employee
                    log.info("Writing employee name to excel")
                    sheet.cell(row=i, column=3).value = model
                    log.info("Writing model name to excel")
                    sheet.cell(row=i, column=4).value =status
                    log.info("Writing status to excel")
                    workbook.save(self.excel_Location)
                    inventorySearchpage.InventoryImeiSearch().clear()


